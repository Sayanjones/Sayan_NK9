"""
Assignment 2: Translation with LLM (English -> Hindi)
Model: facebook/nllb-200-distilled-600M (HuggingFace)
Metrics: BLEU, CHRF, TER
"""

import pandas as pd
from transformers import pipeline, AutoTokenizer, AutoModelForSeq2SeqLM
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sacrebleu
import time

# ── 1. Load cleaned dataset from Assignment 1 ─────────────────────────────────
print("Loading Assignment 1 dataset...")
df_full = pd.read_csv("assignment1_cleaned.csv")
print(f"Total rows available: {len(df_full)}")

# Take exactly 150 English sentences
df = df_full[['English_Sentence', 'Hindi_Sentence']].head(150).copy()
df.reset_index(drop=True, inplace=True)
print(f"Selected 150 sentences for translation.")

# ── 2. Load LLM Translation Model ─────────────────────────────────────────────
# Using facebook/nllb-200-distilled-600M
# NOT Google Translate, NOT Helsinki-NLP — as per assignment rules
print("\nLoading translation model (facebook/nllb-200-distilled-600M)...")
print("This may take a few minutes on first run (downloading ~2.5GB)...")

model_name = "facebook/nllb-200-distilled-600M"
tokenizer = AutoTokenizer.from_pretrained(model_name)
model = AutoModelForSeq2SeqLM.from_pretrained(model_name)

translator = pipeline(
    "translation",
    model=model,
    tokenizer=tokenizer,
    src_lang="eng_Latn",   # English
    tgt_lang="hin_Deva",   # Hindi (Devanagari)
    max_length=512,
    device=-1              # CPU; change to 0 if you have GPU
)

print("Model loaded successfully!")

# ── 3. Translate 150 English sentences ────────────────────────────────────────
print("\nTranslating 150 sentences... (this may take ~5-15 minutes on CPU)")
translated_hindi = []

for i, sentence in enumerate(df['English_Sentence']):
    try:
        result = translator(sentence)
        translation = result[0]['translation_text']
        translated_hindi.append(translation)
    except Exception as e:
        print(f"  Error at row {i}: {e}")
        translated_hindi.append("")

    if (i + 1) % 10 == 0:
        print(f"  Translated {i+1}/150 sentences...")

df['LLM_Hindi_Translation'] = translated_hindi
print("Translation complete!")

# ── 4. Calculate BLEU, CHRF, TER Scores ──────────────────────────────────────
print("\nCalculating evaluation metrics...")

references = df['Hindi_Sentence'].tolist()       # Ground truth Hindi
hypotheses = df['LLM_Hindi_Translation'].tolist() # LLM-generated Hindi

# BLEU Score
bleu = sacrebleu.corpus_bleu(hypotheses, [references])
bleu_score = round(bleu.score, 4)

# CHRF Score
chrf = sacrebleu.corpus_chrf(hypotheses, [references])
chrf_score = round(chrf.score, 4)

# TER Score
ter = sacrebleu.corpus_ter(hypotheses, [references])
ter_score = round(ter.score, 4)

print(f"\n{'='*40}")
print(f"  BLEU Score : {bleu_score}")
print(f"  CHRF Score : {chrf_score}")
print(f"  TER  Score : {ter_score}")
print(f"{'='*40}")

# ── 5. Save scores to scores.txt ──────────────────────────────────────────────
with open("scores.txt", "w", encoding="utf-8") as f:
    f.write("=" * 50 + "\n")
    f.write("  Assignment 2 — Translation Evaluation Scores\n")
    f.write("  Model: facebook/nllb-200-distilled-600M\n")
    f.write("  Task : English → Hindi Translation\n")
    f.write("  Sentences Evaluated: 150\n")
    f.write("=" * 50 + "\n\n")
    f.write(f"BLEU Score : {bleu_score}\n")
    f.write(f"  - Measures n-gram overlap between translation and reference.\n")
    f.write(f"  - Range: 0-100. Higher is better.\n\n")
    f.write(f"CHRF Score : {chrf_score}\n")
    f.write(f"  - Character n-gram F-score. Better for morphologically rich languages like Hindi.\n")
    f.write(f"  - Range: 0-100. Higher is better.\n\n")
    f.write(f"TER Score  : {ter_score}\n")
    f.write(f"  - Translation Edit Rate. Measures edits needed to match reference.\n")
    f.write(f"  - Range: 0+. Lower is better.\n\n")
    f.write("=" * 50 + "\n")
    f.write(f"Full BLEU details:\n{bleu}\n\n")
    f.write(f"Full CHRF details:\n{chrf}\n\n")
    f.write(f"Full TER  details:\n{ter}\n")

print("✅ Saved: scores.txt")

# ── 6. Save Excel Output ──────────────────────────────────────────────────────
output_file = "Assignment2_Translation_Output.xlsx"
wb = Workbook()

# --- Sheet 1: Translations ---
ws1 = wb.active
ws1.title = "Translations"

headers = ['Original English Sentence', 'Model-Generated Hindi Translation', 'Reference Hindi Sentence']

header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill  = PatternFill('solid', start_color='1A5276')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin_border  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

for col_idx, h in enumerate(headers, start=1):
    cell = ws1.cell(row=1, column=col_idx, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = thin_border

fill_light = PatternFill('solid', start_color='D6EAF8')
fill_white  = PatternFill('solid', start_color='FFFFFF')
data_font   = Font(name='Arial', size=10)

for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    fill = fill_light if row_idx % 2 == 0 else fill_white
    values = [row.English_Sentence, row.LLM_Hindi_Translation, row.Hindi_Sentence]
    for col_idx, value in enumerate(values, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = data_font
        cell.fill      = fill
        cell.border    = thin_border
        cell.alignment = left_align

col_widths = [60, 60, 60]
for i, w in enumerate(col_widths, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = 'A2'

# --- Sheet 2: Scores ---
ws2 = wb.create_sheet("Evaluation Scores")

score_headers = ['Metric', 'Score', 'Description', 'Better When']
for col_idx, h in enumerate(score_headers, start=1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = thin_border

score_data = [
    ['BLEU',  bleu_score, 'N-gram precision overlap between hypothesis and reference', 'Higher'],
    ['CHRF',  chrf_score, 'Character n-gram F-score, robust for morphologically rich languages', 'Higher'],
    ['TER',   ter_score,  'Translation Edit Rate — edits needed to match reference', 'Lower'],
]

score_font = Font(name='Arial', size=11, bold=True)
for row_idx, row_data in enumerate(score_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border    = thin_border
        cell.alignment = center_align
        if col_idx == 2:
            cell.font = score_font
            cell.fill = PatternFill('solid', start_color='A9DFBF')

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 65
ws2.column_dimensions['D'].width = 14

ws2.cell(row=6, column=1, value='Model Used:').font   = Font(bold=True, name='Arial')
ws2.cell(row=6, column=2, value='facebook/nllb-200-distilled-600M')
ws2.cell(row=7, column=1, value='Sentences:').font    = Font(bold=True, name='Arial')
ws2.cell(row=7, column=2, value=150)
ws2.cell(row=8, column=1, value='Direction:').font    = Font(bold=True, name='Arial')
ws2.cell(row=8, column=2, value='English → Hindi')

wb.save(output_file)
print(f"✅ Saved: {output_file}")
print("\n🎉 Assignment 2 Complete!")
print(f"   Files generated:")
print(f"   1. {output_file}")
print(f"   2. scores.txt")
