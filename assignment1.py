"""
Assignment 1: English-Hindi Dataset Processing and Analysis
Dataset: ainlpml/english-hindi (HuggingFace)
"""

import pandas as pd
from datasets import load_dataset
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 1. Load Dataset ───────────────────────────────────────────────────────────
print("Loading dataset from HuggingFace...")
ds = load_dataset("ainlpml/english-hindi", use_auth_token=True)
print(ds)

split = list(ds.keys())[0]
df_raw = ds[split].to_pandas()
print(f"\nRaw dataset shape: {df_raw.shape}")
print(f"Columns: {df_raw.columns.tolist()}")
print(df_raw.head(3))

# ── 2. Parse the 'text' column ────────────────────────────────────────────────
# The dataset has a single 'text' column. We need to figure out its structure.
# It may be tab-separated "English\tHindi" or alternating rows, or JSON-like.

sample = str(df_raw['text'].iloc[0])
print(f"\nSample row: {repr(sample)}")

if '\t' in sample:
    # Tab-separated: "English sentence \t Hindi sentence"
    print("Detected: TAB-separated format")
    df_raw[['English_Sentence', 'Hindi_Sentence']] = df_raw['text'].str.split('\t', n=1, expand=True)

elif '|||' in sample:
    # Pipe-separated
    print("Detected: ||| separated format")
    df_raw[['English_Sentence', 'Hindi_Sentence']] = df_raw['text'].str.split('\|\|\|', n=1, expand=True)

elif '\n' in sample:
    # Newline-separated within single field
    print("Detected: newline-separated format")
    df_raw[['English_Sentence', 'Hindi_Sentence']] = df_raw['text'].str.split('\n', n=1, expand=True)

else:
    # Alternating rows: even = English, odd = Hindi
    print("Detected: alternating rows format")
    english_rows = df_raw['text'].iloc[0::2].reset_index(drop=True)
    hindi_rows   = df_raw['text'].iloc[1::2].reset_index(drop=True)
    df_raw = pd.DataFrame({
        'English_Sentence': english_rows,
        'Hindi_Sentence':   hindi_rows
    })

df = df_raw[['English_Sentence', 'Hindi_Sentence']].copy()
df.dropna(inplace=True)
df = df[df['English_Sentence'].str.strip() != '']
df = df[df['Hindi_Sentence'].str.strip() != '']
df['English_Sentence'] = df['English_Sentence'].str.strip()
df['Hindi_Sentence']   = df['Hindi_Sentence'].str.strip()
df.reset_index(drop=True, inplace=True)

print(f"\nAfter parsing: {len(df)} rows")
print(df.head(3))

# ── 3. Word Count ─────────────────────────────────────────────────────────────
df['Word_Count_English'] = df['English_Sentence'].apply(lambda x: len(str(x).split()))
df['Word_Count_Hindi']   = df['Hindi_Sentence'].apply(lambda x: len(str(x).split()))

# ── 4. Filter word count 5–55 in BOTH languages ───────────────────────────────
df = df[
    df['Word_Count_English'].between(5, 55) &
    df['Word_Count_Hindi'].between(5, 55)
].reset_index(drop=True)
print(f"After word-count filter (5–55): {len(df)} rows")

# ── 5. Word Count Difference Filter ±10 ──────────────────────────────────────
df['WC_Difference'] = df['Word_Count_English'] - df['Word_Count_Hindi']
df = df[df['WC_Difference'].between(-10, 10)].reset_index(drop=True)
print(f"After WC-difference filter (±10): {len(df)} rows")

# ── 6. Character Counts ───────────────────────────────────────────────────────
df['Char_Count_English'] = df['English_Sentence'].apply(len)
df['Char_Count_Hindi']   = df['Hindi_Sentence'].apply(len)
df['CC_Difference']      = df['Char_Count_English'] - df['Char_Count_Hindi']

df = df[[
    'English_Sentence', 'Hindi_Sentence',
    'Word_Count_English', 'Word_Count_Hindi', 'WC_Difference',
    'Char_Count_English', 'Char_Count_Hindi', 'CC_Difference'
]]

print(f"\nFinal dataset: {len(df)} rows")
print(df.head(3))

# ── 7. Save to Excel ──────────────────────────────────────────────────────────
output_file = "Assignment1_English_Hindi_Dataset.xlsx"
wb = Workbook()
ws = wb.active
ws.title = "Cleaned Dataset"

headers = [
    'English Sentences', 'Hindi Sentences',
    'Word Count (English)', 'Word Count (Hindi)', 'Difference (Word Count)',
    'Character Count (English)', 'Character Count (Hindi)', 'Difference (Character Count)'
]

header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill  = PatternFill('solid', start_color='2E4057')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin_border  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

fill_light = PatternFill('solid', start_color='EBF5FB')
fill_white  = PatternFill('solid', start_color='FFFFFF')
data_font   = Font(name='Arial', size=10)

for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    fill = fill_light if row_idx % 2 == 0 else fill_white
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = data_font
        cell.fill      = fill
        cell.border    = thin_border
        cell.alignment = left_align if col_idx <= 2 else center_align

col_widths = [50, 50, 22, 20, 25, 25, 23, 28]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A2'

# Summary sheet
ws2 = wb.create_sheet("Summary")
summary_data = [
    ['Metric', 'Value'],
    ['Total rows after all filters', len(df)],
    ['Avg Word Count (English)', round(df['Word_Count_English'].mean(), 2)],
    ['Avg Word Count (Hindi)',   round(df['Word_Count_Hindi'].mean(), 2)],
    ['Avg Char Count (English)', round(df['Char_Count_English'].mean(), 2)],
    ['Avg Char Count (Hindi)',   round(df['Char_Count_Hindi'].mean(), 2)],
    ['Min WC Difference', int(df['WC_Difference'].min())],
    ['Max WC Difference', int(df['WC_Difference'].max())],
]
for r in summary_data:
    ws2.append(r)
ws2['A1'].font = Font(bold=True, name='Arial')
ws2['B1'].font = Font(bold=True, name='Arial')
ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 20

wb.save(output_file)
print(f"\n✅ Saved: {output_file}")

df.to_csv("assignment1_cleaned.csv", index=False, encoding='utf-8-sig')
print("✅ Saved: assignment1_cleaned.csv (use this for Assignment 2)")
